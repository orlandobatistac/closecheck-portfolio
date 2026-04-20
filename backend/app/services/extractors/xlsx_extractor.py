"""
XLSX extractor — openpyxl, each sheet rendered as CSV-like text.
"""
import logging
from pathlib import Path

from app.services.extractors import BaseExtractor, ParsedDocument, compute_sha256

logger = logging.getLogger(__name__)


class XLSXExtractor(BaseExtractor):

    def extract(self, path: Path) -> ParsedDocument:
        sha = compute_sha256(path)
        warnings: list[str] = []

        try:
            import openpyxl
        except ImportError:
            return ParsedDocument(
                filename=path.name,
                file_type="xlsx",
                text="",
                extraction_method="failed",
                warnings=["openpyxl is not installed."],
                sha256=sha,
            )

        try:
            wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        except Exception as exc:  # noqa: BLE001
            return ParsedDocument(
                filename=path.name,
                file_type="xlsx",
                text="",
                extraction_method="failed",
                warnings=[f"Cannot open XLSX file: {exc}"],
                sha256=sha,
            )

        sheets_text: list[str] = []
        sheet_names: list[str] = []

        for sheet_name in wb.sheetnames:
            try:
                ws = wb[sheet_name]
                rows_text: list[str] = []
                row_count = 0

                for row in ws.iter_rows(values_only=True):
                    # Skip entirely empty rows
                    if all(cell is None for cell in row):
                        continue
                    cells = [str(cell) if cell is not None else "" for cell in row]
                    rows_text.append(",".join(cells))
                    row_count += 1

                if rows_text:
                    block = f"--- Sheet: {sheet_name} ---\n" + "\n".join(rows_text)
                    sheets_text.append(block)
                    sheet_names.append(sheet_name)

            except Exception as exc:  # noqa: BLE001
                warnings.append(f"Could not read sheet '{sheet_name}': {exc}")

        wb.close()

        return ParsedDocument(
            filename=path.name,
            file_type="xlsx",
            text="\n\n".join(sheets_text).strip(),
            metadata={"sheet_count": len(sheet_names), "sheets": sheet_names},
            extraction_method="native",
            warnings=warnings,
            sha256=sha,
        )
